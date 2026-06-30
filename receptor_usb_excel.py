
import serial
import serial.tools.list_ports
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import time

# ── Configuración ────────────────────────────────────
BAUDRATE     = 9600
ARCHIVO_EXCEL = "laguna_datos.xlsx"

# ── Detectar puerto automáticamente ─────────────────
def detectar_puerto():
    puertos = serial.tools.list_ports.comports()
    for puerto in puertos:
        # El ESP32-C3 suele aparecer con estos nombres
        if any(x in puerto.description for x in ["USB", "CP210", "CH340", "UART", "ESP"]):
            return puerto.device
    return None

# ── Inicializar Excel ────────────────────────────────
def inicializar_excel():
    if os.path.exists(ARCHIVO_EXCEL):
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
        print(f"Archivo existente encontrado: {ARCHIVO_EXCEL}")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos Laguna"
        ws.append(["Fecha", "Hora", "Distancia (cm)", "Temperatura (C)"])
        # Ancho de columnas
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        wb.save(ARCHIVO_EXCEL)
        print(f"Archivo nuevo creado: {ARCHIVO_EXCEL}")
    return wb, ws

# ── Guardar fila en Excel ────────────────────────────
def guardar_dato(wb, ws, distancia, temperatura):
    ahora = datetime.now()
    ws.append([
        ahora.strftime("%d/%m/%Y"),
        ahora.strftime("%H:%M:%S"),
        float(distancia),
        float(temperatura)
    ])
    wb.save(ARCHIVO_EXCEL)
    print(f"[{ahora.strftime('%H:%M:%S')}] Distancia: {distancia} cm | Temperatura: {temperatura} C")

# ── Main ─────────────────────────────────────────────
def main():
    # Detectar puerto
    puerto = detectar_puerto()

    if puerto is None:
        print("No se detectó el ESP32-C3 automáticamente.")
        print("Puertos disponibles:")
        for p in serial.tools.list_ports.comports():
            print(f"  {p.device} → {p.description}")
        puerto = input("Escribí el puerto manualmente (ej: COM3 o /dev/ttyUSB0): ").strip()

    print(f"Conectando al puerto: {puerto}")

    wb, ws = inicializar_excel()

    try:
        with serial.Serial(puerto, BAUDRATE, timeout=2) as ser:
            print(f"✔ Conectado a {puerto}")
            print(f"Guardando datos en: {ARCHIVO_EXCEL}")
            print("Presioná Ctrl+C para detener.\n")

            # Esperar que el ESP32 se inicialice
            time.sleep(2)
            ser.flushInput()

            while True:
                if ser.in_waiting > 0:
                    linea = ser.readline().decode("utf-8").strip()

                    # Ignorar líneas de encabezado del ESP32
                    if "===" in linea or "Esperando" in linea or linea == "":
                        continue

                    # Procesar formato "distancia,temperatura"
                    partes = linea.split(",")
                    if len(partes) == 2:
                        distancia   = partes[0].strip()
                        temperatura = partes[1].strip()
                        try:
                            guardar_dato(wb, ws, distancia, temperatura)
                        except ValueError:
                            print(f"Dato inválido ignorado: {linea}")
                    else:
                        print(f"Línea ignorada: {linea}")

    except serial.SerialException as e:
        print(f"Error de conexión: {e}")
    except KeyboardInterrupt:
        print("\n✔ Detenido por el usuario.")
        print(f"Datos guardados en: {ARCHIVO_EXCEL}")

if __name__ == "__main__":
    main()
